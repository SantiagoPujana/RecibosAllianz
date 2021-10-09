#!/usr/bin/env python
#_*_ coding: utf-8 _*_

from modules.desingPatterns import Singleton
from modules.logs import writeTXT
from modules.assignedValues import AssignedValues
from modules.colors import Colors
from modules.progressBar import ProgressBar
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

class ExcelFileSettings(metaclass=Singleton):

    def __init__(self):

        self.__excel_file = AssignedValues.getExcelFile()
        self.__report = AssignedValues.getReport()
        self.__receipt_info = AssignedValues.getReceiptInfo()
        self.__GREEN = Colors().green()
        self.__RED = Colors().red()
        self.__red_font = Font(name="Century Gothic", size=11, color="FFFF0000")
        self.__excel_file_object = load_workbook(self.__excel_file)
        self.__excel_sheet_object = self.__excel_file_object[AssignedValues.getExcelSheet()]

    def writeExcelFile(self, reference, datepayday, value, period):

        progress_bar = ProgressBar().getProgressBar()

        font = Font(name="Century Gothic", size=11, color="00000000")
        alignment = Alignment(horizontal="center", vertical="center")
        cell_color = PatternFill(fill_type='solid', fgColor='FFFFFF00')

        self.__excel_sheet_object[self.__receipt_info[3]].value = int(reference)
        self.__excel_sheet_object[self.__receipt_info[4]].value = period
        self.__excel_sheet_object[self.__receipt_info[5]].value = int(value.replace(".",""))
        self.__excel_sheet_object["I" + self.__receipt_info[6][1:]].value = None

        try:
            self.__excel_sheet_object[self.__receipt_info[6]].value = date(
                int(datepayday.split("/")[2]),
                int(datepayday.split("/")[1]),
                int(datepayday.split("/")[0]))

        except ValueError:

            months = {
                'enero' : '01',
                'febrero' : '02',
                'marzo' : '03',
                'abril' : '04',
                'mayo' : '05',
                'junio' : '06',
                'julio' : '07',
                'agosto' : '08',
                'septiembre' : '09',
                'octubre' : '10',
                'noviembre' : '11',
                'diciembre' : '12'
            }

            month = months[datepayday.split("/")[1]]

            self.__excel_sheet_object[self.__receipt_info[6]].value = date(
                int(datepayday.split("/")[2]),
                int(month),
                int(datepayday.split("/")[0]))

        progress_bar.update(5)

        self.__excel_sheet_object[self.__receipt_info[3]].number_format = 'General'
        self.__excel_sheet_object[self.__receipt_info[4]].number_format = 'General'
        self.__excel_sheet_object[self.__receipt_info[5]].number_format = '"$"\\ #,##0'
        self.__excel_sheet_object[self.__receipt_info[6]].number_format = 'mm-dd-yy'
        self.__excel_sheet_object["I" + self.__receipt_info[6][1:]].number_format = 'General'

        progress_bar.update(5)

        for i in range(7):

            if str(type(self.__receipt_info[i])) == "<class 'tuple'>":

                self.__excel_sheet_object[self.__receipt_info[i][1]].fill = cell_color
                self.__excel_sheet_object[self.__receipt_info[i][1]].font = font
                self.__excel_sheet_object[self.__receipt_info[i][1]].alignment = alignment
                self.__excel_sheet_object[self.__receipt_info[i][1]].number_format = 'General'

            else:

                self.__excel_sheet_object[self.__receipt_info[i]].fill = cell_color
                self.__excel_sheet_object[self.__receipt_info[i]].font = font
                self.__excel_sheet_object[self.__receipt_info[i]].alignment = alignment

        progress_bar.update(5)

        progress_bar.write(self.__GREEN + " [+] Los datos se escribieron exitosamente al archivo de Excel.")

        writeTXT(" [+] Los datos se escribieron exitosamente al archivo de Excel.", self.__report)

        self.saveExcelFile()

        progress_bar.update(5)

    def errorProcessExcel(self, text):

        self.__excel_sheet_object["I" + self.__receipt_info[6][1:]].value = text
        self.__excel_sheet_object["I" + self.__receipt_info[6][1:]].font = self.__red_font

        self.saveExcelFile()

    def saveExcelFile(self):

        try:
            self.__excel_file_object.save(filename=self.__excel_file)

        except Exception:

            print(self.__RED + "\n\n [x] Error: Ocurrio un error al guardar el archivo Excel, si lo tienes abierto por favor cierralo o puede que tenga errores.")

            writeTXT("\n\n [x] Error: Ocurrio un error al guardar el archivo Excel, si lo tienes abierto por favor cierralo o puede que tenga errores.", self.__report)

    def getSheetObject(self):
        return self.__excel_sheet_object

    def getCellObject(self, row, column):
        return self.__excel_sheet_object.cell(row=row, column=column)

    def getCellValue(self, row, column):
        return self.__excel_sheet_object.cell(row=row, column=column).value

    def getCellFill(self, row, column):
        return self.__excel_sheet_object.cell(row=row, column=column).fill